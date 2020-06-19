# -----------------------------------------------------------------------------
# Code to automate making offers to MTech students.
#
# May 2020, M. Kaul
# -----------------------------------------------------------------------------
import openpyxl
from dataclasses import dataclass, asdict
from pprint import pprint
from openpyxl.styles import Font
import argparse
import math

# The row from which data starts in master file,
# to skip headers.
MASTER_FILE_ROW_START = 3


@dataclass
class ApplicantRow:
    """A class for holding applicant file row content"""

    a: str
    b: float
    c: str
    d: str
    e: str
    f: str
    g: str
    h: str
    i: float
    j: float
    k: str
    l: str
    m: str
    n: str


@dataclass
class SummaryRow:
    """A class for holding summary file row content"""

    a: str
    b: int
    c: float
    d: float


@dataclass
class OfferRow:
    """A class for holding offer file row content"""

    a: str
    b: str
    c: str
    d: str
    e: str
    f: str
    g: str
    h: str
    i: str
    j: str
    k: float
    l: float
    m: str
    n: str
    o: str
    p: str


@dataclass
class Student:
    """A class for holding Student content"""

    coap_id: str
    gate_score: float
    appl_id: str
    name: str
    gender: str
    category: str
    disabled_flg: str
    gate_id: str
    btech_score: float
    email: str
    mobile: str
    gate_stream: str
    btech_stream: str
    # status: str


@dataclass
class Offer:
    """A class for holding offer content"""

    coap_id: str
    status: str
    reason: str
    name: str
    gender: str
    student_category: str
    offer_seat_category: str
    appl_id: str
    gate_id: str
    disabled_flg: str
    gate_score: float
    btech_score: float
    email: str
    mobile: str
    gate_stream: str
    btech_stream: str


# Remaining seats per category dict
# rem_seats = {"gen": 4, "obc_nc": 3, "ews": 1, "sc": 2, "st": 1, "pwd": 1}
# Factors per category dict
# factors = {"gen": 2, "obc_nc": 2, "ews": 2, "sc": 2, "st": 2, "pwd": 2}
# remaining offers per category dict


def load_all_previous_offers(offers_file, rnd, pos_dict, neg_dict):
    """ load all student offers from each round

    Parameters
    ----------
    offers_file : str
        The name of file to load offer details from
    rnd : int
        The current round
    returns a dict of offer objects
    """
    wb = openpyxl.load_workbook(filename=offers_file)

    # Lets iterate through all worksheets to build our
    # offers dictionary
    offers_dict = {}
    for i, worksheet in enumerate(wb.worksheets, start=1):
        print(f"-- Processing previous worksheet number = {i}")
        # Don't exceed into empty round sheets.
        if i == rnd:
            print(f"time to stop at {i}")
            break

        # Load the rows from file for particular columns of interest
        cols_of_interest = "ABCDEFGHIJKLMNOP"
        rows = [
            OfferRow(
                *(worksheet[f"{column}{row}"].value for column in cols_of_interest)
            )
            for row in range(2, worksheet.max_row + 1)
        ]
        # Iterate through the rows and build up the students
        for r in rows:
            # No need to load students who have been offered in
            # previous round but are still at "Initial_Offer"
            # status.
            if r.c == "Initial_Offer":
                continue

            o = Offer(
                coap_id=r.a,
                status=r.b,
                reason=r.c,
                name=r.d,
                gender=r.e,
                student_category=r.f,
                offer_seat_category=r.g,
                appl_id=r.h,
                gate_id=r.i,
                disabled_flg=r.j,
                gate_score=r.k,
                btech_score=r.l,
                email=r.m,
                mobile=r.n,
                gate_stream=r.o,
                btech_stream=r.p,
            )
            offers_dict[o.coap_id] = {
                "status": o.status,
                "reason": o.reason,
                "name": o.name,
                "gender": o.gender,
                "student_category": o.student_category,
                "offer_seat_category": o.offer_seat_category,
                "appl_id": o.appl_id,
                "gate_id": o.gate_id,
                "disabled_flg": o.disabled_flg,
                "gate_score": o.gate_score,
                "btech_score": o.btech_score,
                "email": o.email,
                "mobile": o.mobile,
                "gate_stream": o.gate_stream,
                "btech_stream": o.btech_stream,
            }

            # Let us just separate out the +ve and -ve COAPs
            if o.status in ["Accept", "Retain"]:
                pos_dict[o.coap_id] = 1
            elif o.status == "Reject":
                neg_dict[o.coap_id] = 1

    return offers_dict


def load_summary(offers_summary_fname, rnd, rem_seats, factors):
    # wb = openpyxl.load_workbook(filename=offers_summary_fname)
    # sheet = wb.sheetnames[rnd - 1]
    # worksheet = wb[sheet]

    wb = openpyxl.load_workbook(filename=offers_summary_fname)
    # Remeber that index is off by one, they start from 0!
    _name = "Round_" + str(rnd)
    worksheet = wb[_name]

    # Load the rows from file for particular columns of interest
    cols_of_interest = "ABCD"
    rows = [
        SummaryRow(*(worksheet[f"{column}{row}"].value for column in cols_of_interest))
        for row in range(2, worksheet.max_row + 1)
    ]
    print(rows)

    for r in rows:
        rem_seats[r.a] = r.b
        factors[r.a] = r.c


def load_students(students_file):
    """ load all student details from this file.

    Parameters
    ----------
    students_file : str
        The name of file to load applicant details from
    returns a list of student objects
    """
    wb = openpyxl.load_workbook(filename=students_file)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]

    # Load the rows from file for particular columns of interest
    # cols_of_interest = ["A", "B", "C", "D", "E", "G", "K", "BM", "BP", "EC"]
    cols_of_interest = "ABCDEFGHIJKLMN"
    rows = [
        ApplicantRow(
            *(worksheet[f"{column}{row}"].value for column in cols_of_interest)
        )
        for row in range(MASTER_FILE_ROW_START, worksheet.max_row + 1)
    ]
    # Iterate through the rows and build up the students
    students = []
    for r in rows:
        # if coap_id == 0, or other single digit strings
        # then lets skip this as its a BTech Application!
        if len(str(r.a)) < 4:
            # print(f'Type of coap_id is = {type(r.a)} and coap_id = {r.a}')
            continue

        # We have to check and merge results from two cols to get
        # btech score
        btech_score = 0.0
        if r.i:
            btech_score = r.i
        else:
            btech_score = r.j

        category = ""
        if r.f == "General/OBC(Creamy layer)":
            category = "gen"
        elif r.f == "OBC(Non Creamy)":
            category = "obc_nc"
        elif r.f == "Economically Weaker Section":
            category = "ews"
        elif r.f == "Scheduled Castes":
            category = "sc"
        elif r.f == "Scheduled Tribes":
            category = "st"

        # This is a PWD
        if r.g == "Yes":
            category = "pwd"

        s = Student(
            coap_id=r.a,
            gate_score=r.b,
            appl_id=r.c,
            name=r.d,
            gender=r.e,
            category=category,
            disabled_flg=r.g,
            gate_id=r.h,
            btech_score=btech_score,
            email=r.k,
            mobile=r.l,
            gate_stream=r.m,
            btech_stream=r.n,
        )
        students.append(s)
        # print(asdict(s))

    # pprint(students)
    return students


def make_offer(s, seat_category, rem_offers, status, reason):
    # s.status = "Offered"
    offers[s.coap_id] = {
        "gate_score": s.gate_score,
        "name": s.name,
        "gender": s.gender,
        "student_category": s.category,
        "disabled_flg": s.disabled_flg,
        "btech_score": s.btech_score,
        "status": status,
        "reason": reason,
        "offer_seat_category": seat_category,  # Want to keep track of seat category
        "gate_id": s.gate_id,
        "appl_id": s.appl_id,
        "email": s.email,
        "mobile": s.mobile,
        "gate_stream": s.gate_stream,
        "btech_stream": s.btech_stream,
    }
    # print(f"for (coap,gate,appl) = ({s.coap_id}, {s.gate_id}, {s.appl_id})")
    # print(offers[(s.coap_id, s.gate_id, s.appl_id)])

    # For "Accept" status, we don't reduce the number of
    # offers.
    if status in ["Retain", "Initial_Offer"]:
     rem_offers[seat_category] -= 1

    return offers


def process_applicants(
    offers, students, rem_offers, pos_dict, neg_dict, prev_offers_dict
):
    """ Process the list of applicants in order and make offers

    Parameters
    ----------
    students : list of Student objects
        The list of students
    returns
    """
    # Iterate through students in desc order of GATE score then
    # btech_score

    print(
        f"pos_dict, neg_dict, prev_offers_dict = {pos_dict}, {neg_dict}, {prev_offers_dict}"
    )

    for s in sorted(
        students, key=lambda x: (x.gate_score, x.btech_score), reverse=True
    ):
        # print(f"\n-- Processing coap_id = {s.coap_id}...")
        # This student was made an offer earlier...
        if s.coap_id in prev_offers_dict:
            print(f"--- Found coap_id in previous offers dict ...")
            # 0. Skip people in negative dict
            if s.coap_id in neg_dict:
                print(
                    f"--- [In -ve list] Found coap_id in previous offers (in sheet) ..."
                )
                continue
            # Make sure to make the offer with SAME seat category and status
            elif s.coap_id in pos_dict:
                print(
                    f"--- [In +ve list] Found coap_id in previous offers (in sheet), so re-offer ..."
                )
                prev_seat_category = prev_offers_dict[s.coap_id]["offer_seat_category"]
                prev_status = prev_offers_dict[s.coap_id]["status"]
                prev_reason = prev_offers_dict[s.coap_id]["reason"]
                offers = make_offer(
                    s, prev_seat_category, rem_offers, prev_status, prev_reason
                )
        # This student was never made an offer by us...
        else:
            # print(f"--- This coap_id was not offered a seat from IITH previously ...")
            # 1. First we fill up general category!
            # print(f"\n---- Processing student category = {s.category}")
            if rem_offers["gen"] > 0:
                # print(f'[before gen offer]: rem_offers --> {rem_offers}')
                offers = make_offer(s, "gen", rem_offers, "Initial_Offer", "")
                # print(f'[after gen offer]: rem_offers --> {rem_offers}')

            # all offers in seat = general category are exhausted.
            elif rem_offers["gen"] <= 0 and s.category != "gen":
                # Check in category other than general
                if rem_offers[s.category] > 0:
                    # print(f'[before category offer]: rem_offers --> {rem_offers}')
                    offers = make_offer(s, s.category, rem_offers, "Initial_Offer", "")
                    # print(f'[after category offer]: rem_offers --> {rem_offers}')
                elif rem_offers[s.category] <= 0:
                    continue

        # If all remaining seats in all categories are
        # down to zero then you can kick out of this
        # iteration early!
        total_cnt = 0
        for category in rem_offers:
            total_cnt += rem_offers[category]
            if total_cnt == 0:
                # print(f"+++ All counts went to zeros = {rem_offers}")
                break


def write_offer_to_workbook(offer_file, offers, rnd):

    wb = openpyxl.load_workbook(filename=offer_file)
    # Remeber that index is off by one, they start from 0!
    _name = "Round_" + str(rnd)
    sh = wb[_name]

    # Column headings
    sh["A1"] = "coap_id"
    sh["B1"] = "status"
    sh["C1"] = "reason"
    sh["D1"] = "name"
    sh["E1"] = "gender"
    sh["F1"] = "student_category"
    sh["G1"] = "offer_seat_category"
    sh["H1"] = "appl_id"
    sh["I1"] = "gate_id"
    sh["J1"] = "disabled_flg"
    sh["K1"] = "gate_score"
    sh["L1"] = "btech_score"
    sh["M1"] = "email"
    sh["N1"] = "mobile"
    sh["O1"] = "gate_stream"
    sh["P1"] = "btech_stream"

    # Just bolding the column headings
    for j in "ABCDEFGHIJKLMNOP":
        sh[j + "1"].font = Font(bold=True)

    # Now start to dump offers dict contents into the worksheet!
    # Note: We start to enumerate from 2 onwards, to skip the
    # column headings.
    for i, (k, v) in enumerate(offers.items(), 2):
        # print(i, k, v)
        sh["A" + str(i)] = k
        sh["B" + str(i)] = v["status"]
        sh["C" + str(i)] = v["reason"]
        sh["D" + str(i)] = v["name"]
        sh["E" + str(i)] = v["gender"]
        sh["F" + str(i)] = v["student_category"]
        sh["G" + str(i)] = v["offer_seat_category"]
        sh["H" + str(i)] = v["appl_id"]
        sh["I" + str(i)] = v["gate_id"]
        sh["J" + str(i)] = v["disabled_flg"]
        sh["K" + str(i)] = v["gate_score"]
        sh["L" + str(i)] = v["btech_score"]
        sh["M" + str(i)] = v["email"]
        sh["N" + str(i)] = v["mobile"]
        sh["O" + str(i)] = v["gate_stream"]
        sh["P" + str(i)] = v["btech_stream"]

    wb.save(filename=offer_file)


def update_cutoffs_in_summary(offers, offers_summary_fname):
    cat_dict = {cat: 99999 for cat in ["gen", "obc_nc", "ews", "sc", "st", "pwd"]}
    wb = openpyxl.load_workbook(filename=offers_summary_fname)
    _name = "Round_" + str(rnd)
    sh = wb[_name]

    # Go through offers and compute cutoffs
    for k, v in offers.items():
        seat_category = v["offer_seat_category"]
        gate_score = v["gate_score"]

        if gate_score < cat_dict[seat_category]:
            cat_dict[seat_category] = gate_score

    # print(f"Category cutoffs --> {cat_dict}")

    sh["D2"] = cat_dict["gen"]
    sh["D3"] = cat_dict["obc_nc"]
    sh["D4"] = cat_dict["ews"]
    sh["D5"] = cat_dict["sc"]
    sh["D6"] = cat_dict["st"]
    sh["D7"] = cat_dict["pwd"]

    wb.save(filename=offers_summary_fname)


#################################################################################
# Main Function
#################################################################################
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-a",
        "--applicants_file",
        type=str,
        required=True,
        help="The master file containing all the applications with coap_id, gate_id, appl_id",
    )
    parser.add_argument(
        "-o",
        "--offers_prefix",
        type=str,
        required=True,
        help="This prefix will use <prefix>_offers.xlsx and <prefix>_summary.xlsx files.",
    )
    parser.add_argument(
        "-r",
        "--round",
        type=int,
        required=True,
        default=1,
        help="Current round of offers to make",
    )
    args = parser.parse_args()
    students_file = args.applicants_file
    offers_prefix = args.offers_prefix
    rnd = args.round

    offers_detail_fname = offers_prefix + "_offers.xlsx"
    offers_summary_fname = offers_prefix + "_summary.xlsx"

    # Dicts we need!
    rem_seats, factors, rem_offers = {}, {}, {}

    load_summary(offers_summary_fname, rnd, rem_seats, factors)
    # Populate rem_offers now!
    for k, v in rem_seats.items():
        #rem_offers[k] = int(v) * int(factors[k])
        rem_offers[k] =  math.ceil( int(v) * float(factors[k]))

    pprint(rem_offers)

    # This will have details of students we made offers to
    offers, prev_offers_dict = {}, {}
    pos_dict, neg_dict = {}, {}
    students = []
    students = load_students(students_file)

    pprint(students)

    if rnd > 1:
        prev_offers_dict = load_all_previous_offers(
            offers_detail_fname, rnd, pos_dict, neg_dict
        )
    # pprint(prev_offers_dict)

    # Process all applications
    process_applicants(
        offers, students, rem_offers, pos_dict, neg_dict, prev_offers_dict
    )
    pprint(offers)

    # Write out the offers to an Excel spreadsheet
    write_offer_to_workbook(offers_detail_fname, offers, rnd)

    # Update cutoffs
    update_cutoffs_in_summary(offers, offers_summary_fname)
