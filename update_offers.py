# -----------------------------------------------------------------------------
# Code to automate to update the status of offers made to MTech students.
#
# May 2020, M. Kaul
#
# File 2: IITH Offered, but accepted OTHER
# File 3: IITH Candidate decision report
# File 4: All IITs consolidated report
# File 6: Coap - Not registered and not responded
#
# -----------------------------------------------------------------------------
import openpyxl
from dataclasses import dataclass, asdict
from pprint import pprint
from openpyxl.styles import Font
import argparse

# The row from which data starts in master file,
# to skip headers.
MASTER_FILE_ROW_START = 3


@dataclass
class ApplicantRow:
    """A class for holding applicant file row content"""

    a: str
    b: float
    c: str
    d: str
    e: str
    f: str
    g: str
    h: str
    i: float
    j: float
    k: str
    l: str
    m: str
    n: str


@dataclass
class OfferRow:
    """A class for holding offer file row content"""

    a: str
    b: str
    c: str
    d: str
    e: str
    f: str
    g: str
    h: str
    i: str
    j: str
    k: float
    l: float
    m: str
    n: str
    o: str
    p: str


@dataclass
class SummaryRow:
    """A class for holding summary file row content"""

    a: str
    b: int
    c: int
    d: float


@dataclass
class UpdateRow:
    """A class for holding updates"""

    coap_id: str
    status: str
    program: str


@dataclass
class Student:
    """A class for holding Student content"""

    coap_id: str
    gate_score: float
    appl_id: str
    name: str
    gender: str
    category: str
    disabled_flg: str
    gate_id: str
    btech_score: float
    email: str
    mobile: str
    gate_stream: str
    btech_stream: str
    #status: str


@dataclass
class Offer:
    """A class for holding offer content"""

    coap_id: str
    status: str
    reason: str
    name: str
    gender: str
    student_category: str
    offer_seat_category: str
    appl_id: str
    gate_id: str
    disabled_flg: str
    gate_score: float
    btech_score: float
    email: str
    mobile: str
    gate_stream: str
    btech_stream: str


def write_updated_summary(offers_summary_fname, rnd, rem_seats, factors):
    wb = openpyxl.load_workbook(filename=offers_summary_fname)
    #_name = "Round_" + str(rnd + 1)
    _name = "Round_" + str(rnd)
    sh = wb[_name]

    # Column headings
    sh["A1"] = "seat_category"
    sh["B1"] = "remaining_seats"
    sh["C1"] = "offers_multiply_by"
    sh["D1"] = "post_round_cutoff"

    # Just bolding the column headings
    for j in "ABCD":
        sh[j + "1"].font = Font(bold=True)

    # Now start to dump seats that remain
    # Note: We start to enumerate from 2 onwards, to skip the
    # column headings.
    for i, (k, v) in enumerate(rem_seats.items(), 2):
        print(i, k, v)
        sh["A" + str(i)] = k
        sh["B" + str(i)] = v
        sh["C" + str(i)] = factors[k]

    wb.save(filename=offers_summary_fname)


def load_summary(offers_summary_fname, rnd, rem_seats, factors):
    # wb = openpyxl.load_workbook(filename=offers_summary_fname)
    # sheet = wb.sheetnames[rnd - 1]
    # worksheet = wb[sheet]

    wb = openpyxl.load_workbook(filename=offers_summary_fname)
    # Remeber that index is off by one, they start from 0!
    _name = "Round_" + str(rnd)
    worksheet = wb[_name]

    # Load the rows from file for particular columns of interest
    cols_of_interest = "ABCD"
    rows = [
        SummaryRow(*(worksheet[f"{column}{row}"].value for column in cols_of_interest))
        for row in range(2, worksheet.max_row + 1)
    ]
    print(rows)

    for r in rows:
        rem_seats[r.a] = r.b
        factors[r.a] = r.c


def load_updates(update_file, coap_id_col, status_col, prog_col):
    """ load all update details from this file.

    Parameters
    ----------
    update_file : str
        The name of file to load applicantion updates to statuses
    returns a list of student objects
    """
    wb = openpyxl.load_workbook(filename=update_file)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]

    # Load the rows from file for particular columns of interest
    # We are only interested in the coap_id and status column in
    # the update file.
    cols_of_interest = [coap_id_col, status_col, prog_col]
    print(
        f"Cols of interest in updates --> {cols_of_interest}, max rows = {worksheet.max_row}"
    )
    rows = [
        UpdateRow(*(worksheet[f"{column}{row}"].value for column in cols_of_interest))
        for row in range(2, worksheet.max_row + 1)
    ]
    #pprint(rows)

    for r in rows:
        r.status = "".join(r.status.split()).lower()

    return rows


def load_students(students_file):
    """ load all student details from this file.

    Parameters
    ----------
    students_file : str
        The name of file to load applicant details from
    returns a list of student objects
    """
    wb = openpyxl.load_workbook(filename=students_file)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]

    # Load the rows from file for particular columns of interest
    #cols_of_interest = ["A", "B", "C", "D", "E", "G", "K", "BM", "BP", "EC"]
    cols_of_interest = "ABCDEFGHIJKLMN"
    rows = [
        ApplicantRow(
            *(worksheet[f"{column}{row}"].value for column in cols_of_interest)
        )
        for row in range(MASTER_FILE_ROW_START, worksheet.max_row + 1)
    ]
    # Iterate through the rows and build up the students
    students_dict = {}
    for r in rows:
        # if coap_id == 0, or other single digit strings
        # then lets skip this as its a BTech Application!
        if len(str(r.a)) < 4:
            # print(f'Type of coap_id is = {type(r.a)} and coap_id = {r.a}')
            continue

        # We have to check and merge results from two cols to get
        # btech score
        btech_score = 0.0
        if r.i:
            btech_score = r.i
        else:
            btech_score = r.j

        category = ""
        if r.f == "General/OBC(Creamy layer)":
            category = "gen"
        elif r.f == "OBC(Non Creamy)":
            category = "obc_nc"
        elif r.f == "Economically Weaker Section":
            category = "ews"
        elif r.f == "Scheduled Castes":
            category = "sc"
        elif r.f == "Scheduled Tribes":
            category = "st"

        # This is a PWD
        if r.g == "Yes":
            category = "pwd"

        s = Student(
            coap_id=r.a,
            gate_score=r.b,
            appl_id=r.c,
            name=r.d,
            gender=r.e,
            category=category,
            disabled_flg=r.g,
            gate_id=r.h,
            btech_score=btech_score,
            email=r.k,
            mobile=r.l,
            gate_stream=r.m,
            btech_stream=r.n
        )
        students_dict[s.coap_id] = {
            "gate_score": s.gate_score,
            "appl_id": s.appl_id,
            "name": s.name,
            "gender": s.gender,
            "category": s.category,
            "disabled_flg": s.disabled_flg,
            "gate_id": s.gate_id,
            "btech_score": s.btech_score,
            "email": s.email,
            "mobile": s.mobile,
            "gate_stream": s.gate_stream,
            "btech_stream": s.btech_stream,
        }
        # print(asdict(s))

    # pprint(students)
    return students_dict


def load_offers(offers_file, rnd):
    """ load all student offers from this file.

    Parameters
    ----------
    offers_file : str
        The name of file to load offer details from
    returns a list of offer objects
    """
    wb = openpyxl.load_workbook(filename=offers_file)
    worksheet = wb["Round_" + str(rnd)]

    # Load the rows from file for particular columns of interest
    cols_of_interest = "ABCDEFGHIJKLMNOP"
    rows = [
        OfferRow(*(worksheet[f"{column}{row}"].value for column in cols_of_interest))
        for row in range(2, worksheet.max_row + 1)
    ]
    # Iterate through the rows and build up the students
    offers_dict = {}
    for r in rows:

        o = Offer(
            coap_id=r.a,
            status=r.b,
            reason=r.c,
            name=r.d,
            gender=r.e,
            student_category=r.f,
            offer_seat_category=r.g,
            appl_id=r.h,
            gate_id=r.i,
            disabled_flg=r.j,
            gate_score=r.k,
            btech_score=r.l,
            email=r.m,
            mobile=r.n,
            gate_stream=r.o,
            btech_stream=r.p
        )
        offers_dict[o.coap_id] = {
            "gate_id": o.gate_id,
            "appl_id": o.appl_id,
            "name": o.name,
            "gender": o.gender,
            "student_category": o.student_category,
            "disabled_flg": o.disabled_flg,
            "gate_score": o.gate_score,
            "btech_score": o.btech_score,
            "offer_seat_category": o.offer_seat_category,
            "status": o.status,
            "reason": o.reason,
            "email": o.email,
            "mobile": o.mobile,
            "gate_stream": o.gate_stream,
            "btech_stream": o.btech_stream
        }
    return offers_dict


def process_updates(
    updates, students_dict, offers_dict, status_map, our_other_flg, rem_seats, program
):
    """ Process the list of updates here.
    Parameters
    ----------
    """
    # Iterate through all the updates
    for up in updates:
        print(f"\n--- update coap_id = {up.coap_id}, status = {up.status}, program={up.program}")

        # If the update program and student program don't match we can skip!
        if up.program not in program:
            print(f"-- Skipped because candidate program = {up.program}")
            continue

        # Found coap_id in the list of applications in master file!
        if up.coap_id in students_dict:
            print(f"+++++++ Found {up.coap_id} in master applications list!")
            # Found him in the current most round of offers
            if up.coap_id in offers_dict:
                print(f"****** Found {up.coap_id} in current offers list!")
                # Which status flag is it ?
                # If it is ours then we must check the
                # statuses carefully to compute remaining seats
                # Let us get what the internal status (int_status) is.
                int_status = status_map[(our_other_flg, up.status)]["status"]
                int_reason = status_map[(our_other_flg, up.status)]["reason"]
                category = offers_dict[up.coap_id]["offer_seat_category"]

                # Lets first check that it is our status
                if our_other_flg == "our":
                    # If we have offered this student and he
                    # accepts or retains then we must reduce
                    # the number of available seats in the
                    # seat_category.
                    if int_status in ["Accept", "Retain"]:
                        print(f'\n------->>>>> [{up.coap_id}] For int_status = {int_status}, we reduce in {category} seats!\n')
                        rem_seats[category] -= 1

                # Stamp the right status and reason on offer
                offers_dict[up.coap_id]["status"] = int_status
                offers_dict[up.coap_id]["reason"] = int_reason

            # coap_id is not in offers file for round, this is
            # an applicant that we did not offer but
            # took up some other inst. offer.
            else:
                # We must insert a new row in the offers dict then!
                o = Offer(
                    coap_id=up.coap_id,
                    gate_id=students_dict[up.coap_id]["gate_id"],
                    appl_id=students_dict[up.coap_id]["appl_id"],
                    name=students_dict[up.coap_id]["name"],
                    gender=students_dict[up.coap_id]["gender"],
                    student_category=students_dict[up.coap_id]["category"],
                    disabled_flg=students_dict[up.coap_id]["disabled_flg"],
                    gate_score=students_dict[up.coap_id]["gate_score"],
                    btech_score=students_dict[up.coap_id]["btech_score"],
                    email=students_dict[up.coap_id]["email"],
                    mobile=students_dict[up.coap_id]["mobile"],
                    gate_stream=students_dict[up.coap_id]["gate_stream"],
                    btech_stream=students_dict[up.coap_id]["btech_stream"],
                    offer_seat_category="",
                    status="Reject",
                    reason="IITH never offered, accepted other offer",
                )
                offers_dict[o.coap_id] = {
                    "gate_id": o.gate_id,
                    "appl_id": o.appl_id,
                    "name": o.name,
                    "gender": o.gender,
                    "student_category": o.student_category,
                    "disabled_flg": o.disabled_flg,
                    "gate_score": o.gate_score,
                    "btech_score": o.btech_score,
                    "offer_seat_category": o.offer_seat_category,
                    "status": o.status,
                    "reason": o.reason,
                    "email": o.email,
                    "mobile": o.mobile,
                    "gate_stream": o.gate_stream,
                    "btech_stream": o.btech_stream
                }

    return offers_dict


def write_updated_offers_to_workbook(offer_file, offers_dict, rnd):

    wb = openpyxl.load_workbook(filename=offer_file)
    sh = wb["Round_" + str(rnd)]

    # Column headings
    sh["A1"] = "coap_id"
    sh["B1"] = "status"
    sh["C1"] = "reason"
    sh["D1"] = "name"
    sh["E1"] = "gender"
    sh["F1"] = "student_category"
    sh["G1"] = "offer_seat_category"
    sh["H1"] = "appl_id"
    sh["I1"] = "gate_id"
    sh["J1"] = "disabled_flg"
    sh["K1"] = "gate_score"
    sh["L1"] = "btech_score"
    sh["M1"] = "email"
    sh["N1"] = "mobile"
    sh["O1"] = "gate_stream"
    sh["P1"] = "btech_stream"

    # Just bolding the column headings
    for j in "ABCDEFGHIJKLMNOP":
        sh[j + "1"].font = Font(bold=True)

    # Now start to dump offers dict contents into the worksheet!
    # Note: We start to enumerate from 2 onwards, to skip the
    # column headings.
    for i, (k, v) in enumerate(offers_dict.items(), 2):
        # print(i, k, v)
        sh["A" + str(i)] = k
        sh["B" + str(i)] = v["status"]
        sh["C" + str(i)] = v["reason"]
        sh["D" + str(i)] = v["name"]
        sh["E" + str(i)] = v["gender"]
        sh["F" + str(i)] = v["student_category"]
        sh["G" + str(i)] = v["offer_seat_category"]
        sh["H" + str(i)] = v["appl_id"]
        sh["I" + str(i)] = v["gate_id"]
        sh["J" + str(i)] = v["disabled_flg"]
        sh["K" + str(i)] = v["gate_score"]
        sh["L" + str(i)] = v["btech_score"]
        sh["M" + str(i)] = v["email"]
        sh["N" + str(i)] = v["mobile"]
        sh["O" + str(i)] = v["gate_stream"]
        sh["P" + str(i)] = v["btech_stream"]

    wb.save(filename=offer_file)


#################################################################################
# Main Function
#################################################################################
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    # This is done to ensure that you can only pass at a time either the
    # "our_status_col" or "other_status_col" so we pick the status from
    # there.
    group = parser.add_mutually_exclusive_group()

    parser.add_argument(
        "-a",
        "--applicants_file",
        type=str,
        required=True,
        help="The master file containing all the applications with coap_id, gate_id, appl_id",
    )
    parser.add_argument(
        "-u",
        "--update_file",
        type=str,
        required=True,
        help="The coap updates file containing all the application status changes",
    )
    parser.add_argument(
        "-c",
        "--coap_id_col",
        type=str,
        required=True,
        help="The column name in updates spreadsheet where coap_id is present. E.g. A means look in column A",
    )
    parser.add_argument(
        "-op",
        "--offers_prefix",
        type=str,
        required=True,
        help="This prefix will use <prefix>_offers.xlsx and <prefix>_summary.xlsx files.",
    )
    parser.add_argument(
        "-prg",
        "--program",
        type=str,
        required=True,
        help="This is the program offered (e.g. CSE, NIS)",
    )
    parser.add_argument(
        "-pcol",
        "--program_col",
        type=str,
        required=True,
        help="This is the column name in updates spreadsheet where program offered is present.",
    )
    parser.add_argument(
        "-r",
        "--round",
        type=int,
        required=True,
        default=1,
        help="Current round of offers to update",
    )
    group.add_argument(
        "-our",
        "--our_status_col",
        type=str,
        help="The column in updates spreadsheet where iith_status is present.",
    )
    group.add_argument(
        "-oth",
        "--other_status_col",
        type=str,
        help="The column in updates spreadsheet where other_status is present.",
    )

    args = parser.parse_args()

    students_file = args.applicants_file
    update_file = args.update_file
    offers_prefix = args.offers_prefix
    coap_id_col = args.coap_id_col
    our_status_col = args.our_status_col
    other_status_col = args.other_status_col
    program = args.program
    prog_col = args.program_col
    rnd = args.round

    offers_detail_fname = offers_prefix + "_offers.xlsx"
    offers_summary_fname = offers_prefix + "_summary.xlsx"

    # Dicts we need!
    rem_seats, factors, rem_offers = {}, {}, {}

    load_summary(offers_summary_fname, rnd, rem_seats, factors)
    # Populate rem_offers now!
    for k, v in rem_seats.items():
        rem_offers[k] = int(v) * int(factors[k])

    # pprint(rem_offers)

    # This will have details of students we made offers to
    offers_dict = {}
    students_dict = {}
    students_dict = load_students(students_file)
    # pprint(students_dict)

    offers_dict = load_offers(offers_detail_fname, rnd)
    # pprint(offers_dict)

    status_col = ""
    if not our_status_col:
        status_col = other_status_col
    else:
        status_col = our_status_col

    # Now let us load the updates!
    updates = []
    updates_list = load_updates(update_file, coap_id_col, status_col, prog_col)

    #
    # Let us make a status map, from what is in the update file to
    # our internal offer statuses and reasons
    #
    status_map = {
        ("our", "acceptandfreeze"): {
            "status": "Accept",
            "reason": "IITH offered, accepted our offer",
        },
        ("our", "rejectandwait"): {
            "status": "Reject",
            "reason": "IITH offered, rejected our offer",
        },
        ("our", "retainandwait"): {
            "status": "Retain",
            "reason": "IITH offered, retained our offer",
        },
        ("oth", "acceptandfreeze"): {
            "status": "Reject",
            "reason": "IITH offered, accepted other offer",
        },
    }
    # pprint(status_map)

    our_other_flg = ""
    if not our_status_col:
        our_other_flg = "oth"
    else:
        our_other_flg = "our"

    updated_offers_dict = {}

    updated_offers_dict = process_updates(
        updates_list, students_dict, offers_dict, status_map, our_other_flg, rem_seats, program
    )
    # print(f'After processing updates: {updated_offers_dict}')
    pprint(updated_offers_dict)
    # Write out the latest offers
    write_updated_offers_to_workbook(offers_detail_fname, updated_offers_dict, rnd)
    # Update the remaining seats too in the summary file
    write_updated_summary(offers_summary_fname, rnd, rem_seats, factors)
